import sys
import psycopg2
import pandas as pd
from datetime import date, timedelta, datetime
import smtplib
from email.mime.text import MIMEText
import ssl
from tabulate import tabulate
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os

# --- FUNCIONES DE FORMATEO ---
def formatear_fechas_para_tabla(fechas_lista):
    """Convierte una lista de fechas en formato YYYY-MM-DD a tabla HTML con Año, Mes y Días"""
    from datetime import datetime
    from collections import defaultdict
    
    # Agrupar fechas por año y mes
    fechas_agrupadas = defaultdict(lambda: defaultdict(list))
    
    for fecha_str in sorted(set(fechas_lista)):
        try:
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d')
            año = fecha.year
            mes = fecha.month
            dia = fecha.day
            fechas_agrupadas[año][mes].append(dia)
        except:
            continue
    
    # Generar HTML de la tabla
    filas_html = []
    for año in sorted(fechas_agrupadas.keys()):
        for mes in sorted(fechas_agrupadas[año].keys()):
            dias = sorted(fechas_agrupadas[año][mes])
            dias_str = ' '.join([f"{dia:02d}" for dia in dias])
            fila = f"""
                <tr>
                    <td style="border: 1px solid #ddd; padding: 8px;">{año}</td>
                    <td style="border: 1px solid #ddd; padding: 8px;">{mes:02d}</td>
                    <td style="border: 1px solid #ddd; padding: 8px;">{dias_str}</td>
                </tr>
            """
            filas_html.append(fila)
    
    return ''.join(filas_html)

def generar_tabla_resumen_html(resumen_eots):
    """Genera una tabla HTML del resumen de EOTs con incumplimientos"""
    tabla_html = """
    <table style="border-collapse: collapse; width: 100%; margin: 10px 0; font-size: 12px;">
        <thead>
            <tr style="background-color: #f2f2f2;">
                <th style="border: 1px solid #ddd; padding: 8px; text-align: left;">Nombre EOT</th>
                <th style="border: 1px solid #ddd; padding: 8px; text-align: center;">Días con <100%</th>
                <th style="border: 1px solid #ddd; padding: 8px; text-align: center;">Promedio % Ruta</th>
                <th style="border: 1px solid #ddd; padding: 8px; text-align: center;">Buses Reportado</th>
                <th style="border: 1px solid #ddd; padding: 8px; text-align: center;">Buses en Ruta</th>
                <th style="border: 1px solid #ddd; padding: 8px; text-align: left;">Fechas con Problemas</th>
            </tr>
        </thead>
        <tbody>
    """
    
    for idx, row in resumen_eots.iterrows():
        # Determinar color de fila según el número de días con problemas
        color_fila = "#ffebee" if row['Días con <100%'] >= 3 else "#fff3e0" if row['Días con <100%'] >= 2 else "#f3e5f5"
        
        fila = f"""
            <tr style="background-color: {color_fila};">
                <td style="border: 1px solid #ddd; padding: 8px; font-weight: bold;">{idx}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{row['Días con <100%']}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{row['Promedio % Ruta']:.2f}%</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{row['Buses Reportado']:.0f}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{row['Buses en Ruta']:.0f}</td>
                <td style="border: 1px solid #ddd; padding: 8px; font-size: 11px;">{row['Fechas con problemas']}</td>
            </tr>
        """
        tabla_html += fila
    
    tabla_html += """
        </tbody>
    </table>
    """
    
    return tabla_html

# --- FUNCIONES DE BASE DE DATOS ---
def crear_tabla_detalles_incumplimientos():
    """Crea la tabla de detalles de incumplimientos si no existe"""
    try:
        connCID = psycopg2.connect(host=hostCID, database=databaseCID, user=userCID, password=passwordCID, port=portCID)
        cur = connCID.cursor()
        
        # Crear tabla de detalles de incumplimientos
        query_crear_tabla = """
            CREATE TABLE IF NOT EXISTS verificacion_res_65.detalles_incumplimientos (
                id SERIAL PRIMARY KEY,
                eot_nombre VARCHAR(255) NOT NULL,
                cod_catalogo VARCHAR(50) NOT NULL,
                id_eot_vmt_hex VARCHAR(50) NOT NULL,
                e_mail VARCHAR(255),
                fecha_incumplimiento DATE NOT NULL,
                buses_reportado INTEGER NOT NULL,
                buses_operando_ruta INTEGER NOT NULL,
                porcentaje_cumplimiento DECIMAL(5,2) NOT NULL,
                fecha_hora_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                nro_notificacion INTEGER,
                nivel INTEGER DEFAULT 2,
                observaciones TEXT
            );
        """
        
        cur.execute(query_crear_tabla)
        connCID.commit()
        
        # Crear índices para mejorar el rendimiento
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_detalles_eot 
            ON verificacion_res_65.detalles_incumplimientos(eot_nombre, cod_catalogo);
        """)
        
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_detalles_fecha 
            ON verificacion_res_65.detalles_incumplimientos(fecha_incumplimiento);
        """)
        
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_detalles_notificacion 
            ON verificacion_res_65.detalles_incumplimientos(nro_notificacion);
        """)
        
        connCID.commit()
        cur.close()
        connCID.close()
        
        print("✅ Tabla 'detalles_incumplimientos' creada exitosamente")
        return True
        
    except Exception as e:
        print(f"❌ Error al crear tabla detalles_incumplimientos: {e}")
        return False

def insertar_detalles_incumplimientos(resultado_final):
    """Inserta los detalles individuales de incumplimientos en la tabla detalles_incumplimientos"""
    try:
        connCID = psycopg2.connect(host=hostCID, database=databaseCID, user=userCID, password=passwordCID, port=portCID)
        cur = connCID.cursor()
        
        # Preparar datos para inserción individual
        datos_insercion = []
        nivel = 2
        
        for idx, row in resultado_final.iterrows():
            datos_insercion.append((
                row['Nombre EOT'],
                row['Cod Catalogo'],
                row['ID EOT HEX'],
                row['Email'],
                row['Fecha'],
                int(row['Buses Reportado']),
                int(row['Buses operando en ruta']),
                row['Porcentaje_Num'],
                datetime.now(),
                None,  # nro_notificacion será NULL inicialmente
                nivel,
                f"Incumplimiento detectado: {row['Porcentaje_Num']:.2f}% de cumplimiento"
            ))
        
        # Insertar en la tabla detalles_incumplimientos
        query_insert = """
            INSERT INTO verificacion_res_65.detalles_incumplimientos 
            (eot_nombre, cod_catalogo, id_eot_vmt_hex, e_mail, fecha_incumplimiento, 
             buses_reportado, buses_operando_ruta, porcentaje_cumplimiento, 
             fecha_hora_registro, nro_notificacion, nivel, observaciones)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        cur.executemany(query_insert, datos_insercion)
        connCID.commit()
        
        print(f"✅ Se insertaron {len(datos_insercion)} registros detallados de incumplimientos")
        for row in resultado_final.itertuples():
            print(f"  - {row[1]} ({row[5]}): {row[6]:.2f}% - {row[2]} buses reportados, {row[3]} operando")
        
        cur.close()
        connCID.close()
        return True
        
    except Exception as e:
        print(f"❌ Error al insertar detalles de incumplimientos: {e}")
        return False

def insertar_incumplimientos(resultado_final, fecha_inicio, fecha_fin):
    """Inserta los incumplimientos agrupados por EOT en la tabla verificacion_res_65.incumplimientos"""
    try:
        connCID = psycopg2.connect(host=hostCID, database=databaseCID, user=userCID, password=passwordCID, port=portCID)
        cur = connCID.cursor()
        
        # Agrupar por EOT y obtener todas las fechas de incumplimiento
        incumplimientos_agrupados = resultado_final.groupby(['Nombre EOT', 'Cod Catalogo', 'ID EOT HEX', 'Email']).agg({
            'Fecha': lambda x: ', '.join(sorted(x.unique())),  # Todas las fechas concatenadas
            'Buses Reportado': 'mean',
            'Buses operando en ruta': 'mean',
            'Porcentaje_Num': 'mean'
        }).reset_index()
        
        # Preparar datos para inserción
        datos_insercion = []
        nivel = 2  # Mover nivel antes del bucle
        for idx, row in incumplimientos_agrupados.iterrows():
            datos_insercion.append((
                row['Nombre EOT'],
                row['Cod Catalogo'],
                row['ID EOT HEX'],
                row['Email'],
                None,  # nro_notificacion será NULL inicialmente
                fecha_inicio,
                fecha_fin,
                datetime.now(),  # fecha_hora actual
                row['Fecha'],  # todas las fechas de incumplimiento concatenadas
                nivel  # Agregar nivel al final
            ))
        # Insertar en la tabla incumplimientos
        query_insert = """
            INSERT INTO verificacion_res_65.notificaciones 
            (eot_nombre, cod_catalogo, id_eot_vmt_hex, e_mail, nro_notificacion, 
             fecha_inicio, fecha_fin, fecha_hora, fecha_sin_datos, nivel)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        cur.executemany(query_insert, datos_insercion)
        connCID.commit()
        
        print(f"✅ Se insertaron {len(datos_insercion)} incumplimientos agrupados en la base de datos")
        for row in incumplimientos_agrupados.itertuples():
            print(f"  - {row[1]}: {row[5]} (fechas: {row[5]})")
        
        cur.close()
        connCID.close()
        return True
        
    except Exception as e:
        print(f"❌ Error al insertar incumplimientos: {e}")
        return False

def consultar_detalles_incumplimientos(fecha_inicio=None, fecha_fin=None, eot_nombre=None):
    """Consulta los detalles de incumplimientos desde la base de datos"""
    try:
        connCID = psycopg2.connect(host=hostCID, database=databaseCID, user=userCID, password=passwordCID, port=portCID)
        cur = connCID.cursor()
        
        # Construir query dinámico
        query = """
            SELECT id, eot_nombre, cod_catalogo, fecha_incumplimiento, 
                   buses_reportado, buses_operando_ruta, porcentaje_cumplimiento,
                   fecha_hora_registro, nro_notificacion, observaciones
            FROM verificacion_res_65.detalles_incumplimientos
            WHERE 1=1
        """
        params = []
        
        if fecha_inicio:
            query += " AND fecha_incumplimiento >= %s"
            params.append(fecha_inicio)
        
        if fecha_fin:
            query += " AND fecha_incumplimiento <= %s"
            params.append(fecha_fin)
        
        if eot_nombre:
            query += " AND eot_nombre ILIKE %s"
            params.append(f"%{eot_nombre}%")
        
        query += " ORDER BY fecha_incumplimiento DESC, eot_nombre"
        
        cur.execute(query, params)
        resultados = cur.fetchall()
        
        # Convertir a DataFrame
        columnas = ['ID', 'EOT', 'Cod Catalogo', 'Fecha', 'Buses Reportado', 
                   'Buses Operando', 'Porcentaje', 'Fecha Registro', 'Nro Notificación', 'Observaciones']
        
        df_resultados = pd.DataFrame(resultados, columns=columnas)
        
        cur.close()
        connCID.close()
        
        print(f"✅ Se encontraron {len(df_resultados)} registros de detalles de incumplimientos")
        return df_resultados
        
    except Exception as e:
        print(f"❌ Error al consultar detalles de incumplimientos: {e}")
        return pd.DataFrame()

def generar_notificaciones_incumplimientos():
    """Genera notificaciones para EOTs con incumplimientos sin notificar"""
    try:
        connCID = psycopg2.connect(host=hostCID, database=databaseCID, user=userCID, password=passwordCID, port=portCID)
        cur = connCID.cursor()
        
        # Consultar EOTs con notificaciones sin notificar
        query_consulta = """
            SELECT eot_nombre, cod_catalogo, id_eot_vmt_hex, COUNT(*) as cantidad
            FROM verificacion_res_65.notificaciones
            WHERE nro_notificacion IS NULL
              AND fecha_hora >= CURRENT_DATE - INTERVAL '90 days'
            GROUP BY eot_nombre, cod_catalogo, id_eot_vmt_hex
            HAVING COUNT(*) >= 1
        """
        
        cur.execute(query_consulta)
        resultados = cur.fetchall()
        
        if not resultados:
            print("No hay EOTs con incumplimientos pendientes de notificación")
            return []
        
        print(f"📊 Se encontraron {len(resultados)} EOTs con incumplimientos pendientes")
        
        # Crear contenido del correo para empresas con incumplimientos
        notificaciones_generadas = []
        
        for eot_nombre, cod_catalogo, id_eot_vmt_hex, cantidad in resultados:
            # Obtener fechas específicas de incumplimiento para esta EOT
            query_fechas = """
                SELECT fecha_sin_datos
                FROM verificacion_res_65.notificaciones
                WHERE nro_notificacion IS NULL
                  AND cod_catalogo = %s
                  AND id_eot_vmt_hex = %s
                  AND fecha_hora >= CURRENT_DATE - INTERVAL '90 days'
                ORDER BY fecha_sin_datos
            """
            
            cur.execute(query_fechas, (cod_catalogo, id_eot_vmt_hex))
            fechas_incumplimiento = [row[0] for row in cur.fetchall()]
            fechas_str = ', '.join(fechas_incumplimiento)
            nivel = 2
            
            # Insertar en la tabla multas y obtener el ID
            query_infraccion = """
                INSERT INTO verificacion_res_65.multas 
                (eot_nombre, cod_catalogo, id_eot_vmt_hex, fecha_multa, nivel)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id_multa
            """
            
            cur.execute(query_infraccion, (eot_nombre, cod_catalogo, id_eot_vmt_hex, datetime.now(), nivel))
            id_infraccion = cur.fetchone()[0]
            
            # Actualizar las notificaciones con el nro_notificacion correspondiente
            query_update = """
                UPDATE verificacion_res_65.notificaciones
                SET nro_notificacion = %s
                WHERE nro_notificacion IS NULL
                  AND cod_catalogo = %s
                  AND id_eot_vmt_hex = %s
                  AND fecha_hora >= CURRENT_DATE - INTERVAL '90 days'
            """
            
            cur.execute(query_update, (id_infraccion, cod_catalogo, id_eot_vmt_hex))
            
            notificaciones_generadas.append({
                'eot_nombre': eot_nombre,
                'cod_catalogo': cod_catalogo,
                'id_eot_vmt_hex': id_eot_vmt_hex,
                'cantidad': cantidad,
                'fecha_sin_datos': fechas_str,
                'nro_notificacion': id_infraccion
            })
            
            print(f"✅ Notificación generada para {eot_nombre} (ID: {id_infraccion}) - {cantidad} incumplimientos")
        
        connCID.commit()
        cur.close()
        connCID.close()
        
        return notificaciones_generadas
        
    except Exception as e:
        print(f"❌ Error al generar notificaciones de incumplimientos: {e}")
        return []

# --- FUNCIONES DE NOTIFICACIÓN ---
def enviar_email_con_adjunto(destinatario, asunto, cuerpo, archivo_adjunto, remitente, password, servidor_smtp, puerto_smtp):
    mensaje = MIMEMultipart()
    mensaje['From'] = remitente
    mensaje['To'] = destinatario
    mensaje['Subject'] = asunto

    mensaje.attach(MIMEText(cuerpo, 'plain'))

    # Adjuntar PDF
    with open(archivo_adjunto, "rb") as adjunto:
        parte = MIMEBase('application', 'octet-stream')
        parte.set_payload(adjunto.read())
        encoders.encode_base64(parte)
        parte.add_header('Content-Disposition', f'attachment; filename={os.path.basename(archivo_adjunto)}')
        mensaje.attach(parte)

    contexto = ssl.create_default_context()
    with smtplib.SMTP_SSL(servidor_smtp, puerto_smtp, context=contexto) as servidor:
        servidor.login(remitente, password)
        servidor.sendmail(remitente, destinatario, mensaje.as_string())
    print(f"Correo enviado a {destinatario} con el archivo adjunto.")

def enviar_notificacion_incumplimiento(destinatario, nombre_eot, fecha, buses_reportado, buses_operando, porcentaje):
    """Envía notificación de incumplimiento a una empresa específica"""
    asunto = f"ALERTA: Incumplimiento en reporte GPS - {nombre_eot} - {fecha}"
    
    cuerpo = f"""
Estimado/a representante de {nombre_eot},

Le informamos que se ha detectado un incumplimiento en el reporte de datos GPS para la fecha {fecha}:

DETALLES DEL INCUMPLIMIENTO:
- Empresa: {nombre_eot}
- Fecha: {fecha}
- Buses reportados vía GPS: {buses_reportado}
- Buses operando según SNBE: {buses_operando}
- Porcentaje de cumplimiento: {porcentaje}

ACCIONES REQUERIDAS:
1. Verificar el funcionamiento de los dispositivos GPS
2. Asegurar que todos los buses en operación estén reportando datos
3. Contactar al soporte técnico si persisten los problemas

Este reporte es generado automáticamente por el Sistema de Monitoreo de la VMT.

Atentamente,
Sistema de Monitoreo VMT
Viceministerio de Transporte
"""
    
    try:
        # Enviar solo el texto, sin PDF
        mensaje = MIMEText(cuerpo, 'html')
        mensaje['From'] = "billetajevmt@gmail.com"
        mensaje['To'] = destinatario
        mensaje['Subject'] = asunto
        # Agregar copia a vmtdeveloperpy@gmail.com
        mensaje['Cc'] = "vmtdeveloperpy@gmail.com"

        contexto = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=contexto) as servidor:
            servidor.login("billetajevmt@gmail.com", "qlju dhxo jbon exlg")
            # Enviar a destinatario principal y copia
            destinatarios = [destinatario, "vmtdeveloperpy@gmail.com"]
            servidor.sendmail("billetajevmt@gmail.com", destinatarios, mensaje.as_string())
        print(f"Correo enviado a {destinatario} con copia a vmtdeveloperpy@gmail.com")
        return True
    except Exception as e:
        print(f"Error al enviar notificación a {destinatario}: {e}")
        return False

def notificar_incumplimientos_automatico(eots_bajo_100, fecha):
    """Envía notificaciones automáticas a todas las empresas con incumplimientos"""
    notificaciones_enviadas = 0
    notificaciones_fallidas = 0
    
    print(f"\n=== ENVIANDO NOTIFICACIONES DE INCUMPLIMIENTO PARA {fecha} ===")
    
    for idx, row in eots_bajo_100.iterrows():
        nombre_eot = row['Nombre EOT']
        email = row['Email']
        buses_reportado = row['Buses Reportado']
        buses_operando = row['Buses operando en ruta']
        porcentaje = row['Porcentaje en ruta']
        
        # Verificar que el email no esté vacío
        if pd.notna(email) and email.strip():
            print(f"Enviando notificación a {nombre_eot} ({email})...")
            
            if enviar_notificacion_incumplimiento(
                destinatario=email,
                nombre_eot=nombre_eot,
                fecha=fecha,
                buses_reportado=buses_reportado,
                buses_operando=buses_operando,
                porcentaje=porcentaje
            ):
                notificaciones_enviadas += 1
            else:
                notificaciones_fallidas += 1
        else:
            print(f"⚠️  No se puede enviar notificación a {nombre_eot}: Email no disponible")
            notificaciones_fallidas += 1
    
    print(f"\n📊 RESUMEN DE NOTIFICACIONES:")
    print(f"✅ Notificaciones enviadas exitosamente: {notificaciones_enviadas}")
    print(f"❌ Notificaciones fallidas: {notificaciones_fallidas}")
    print(f"📧 Total de empresas notificadas: {notificaciones_enviadas + notificaciones_fallidas}")
    
    return notificaciones_enviadas, notificaciones_fallidas

def enviar_notificacion_resumen_periodo(destinatario, nombre_eot, resumen_eot, fecha_inicio, fecha_fin, fechas_infraccion):
    """Envía notificación de resumen de período con todas las fechas de infracción"""
    asunto = f"NOTIFICACION DE DETECCION DE INCUMPLIMIENTOS - {nombre_eot} - Período {fecha_inicio} a {fecha_fin}"

    cuerpo = f"""
    <html>
    <body style="font-family: Arial, sans-serif; color: #333; line-height: 1.5;">
        <p>Estimado/a representante de <strong>{nombre_eot}</strong>,</p>

        <p>
        Le informamos el resumen de incumplimientos detectados en el reporte de datos GPS 
        para el período <strong>{fecha_inicio} a {fecha_fin}</strong>:
        </p>

        <h3 style="color: #004080;">RESUMEN GENERAL:</h3>
        <ul>
        <li><strong>Empresa:</strong> {nombre_eot}</li>
        <li><strong>Período evaluado:</strong> {fecha_inicio} a {fecha_fin}</li>
        <li><strong>Días con incumplimientos:</strong> {resumen_eot.get('Días con <100%', 'N/A')}</li>
        <li><strong>Promedio de cumplimiento:</strong> {resumen_eot.get('Promedio % Ruta', 'N/A')}%</li>
        <li><strong>Promedio de buses reportados:</strong> {resumen_eot.get('Buses Reportado', 'N/A')}</li>
        <li><strong>Promedio de buses en ruta:</strong> {resumen_eot.get('Buses en Ruta', 'N/A')}</li>
        </ul>

        <h3 style="color: #800000;">FECHAS CON INCUMPLIMIENTOS:</h3>
        <table style="border-collapse: collapse; width: 50%; margin: 10px 0;">
            <thead>
                <tr style="background-color: #f2f2f2;">
                    <th style="border: 1px solid #ddd; padding: 8px; text-align: left;">Año</th>
                    <th style="border: 1px solid #ddd; padding: 8px; text-align: left;">Mes</th>
                    <th style="border: 1px solid #ddd; padding: 8px; text-align: left;">Días</th>
                </tr>
            </thead>
            <tbody>
                {fechas_infraccion}
            </tbody>
        </table>

        <p>Atendiendo lo estabelcido en el <strong>Articulo 1º de la Resolucion GVMT Nº 65/2024</strong> y la <strong>Circulae DMT Nº 05/2025</strong>, 
        se notifica por este medio el incumplimiento en el envío de datos. Atendiendo a dicho incumplimiento se informa que se procederá según 
        lo dispuesto en el <strong>Articulo 5º de la misma Resolución</strong> y la <strong>Circular mencionada</strong>.</p>
        
        <p>En consecuencia, se ha <strong>generado una notificación formal de incumplimiento</strong>, la cual conlleva la aplicación de una multa administrativa conforme al régimen vigente.</p>

        <p><strong>Aclaración importante:</strong> La presente notificación ha sido <strong>generada de manera automática</strong> y, de constatarse 
        la veracidad de las detecciones luego de un análisis más exhaustivo, <strong>podría acarrear sanciones según la Circular N° 5</strong>.</p>

        <p>Asimismo, en caso de que la <strong>Empresa Operadora de Transporte (EOT)</strong> o su respectiva <strong>Empresa Emisora de Mensajes (EEM)</strong> 
        constate que la presente notificación <strong>no corresponda</strong>, podrá responder a este mismo correo, de modo a que nuestro equipo pueda 
        <strong>verificar las discrepancias</strong> que pudiesen existir.</p>
    
        <p>Sin otro particular, saludamos atentamente.</p>

        <div class="footer">
            <p>
                <strong>Coordinación de Innovación y Desarrollo<br>
                Viceministerio de Transporte</strong>
            </p>
        </div>
    </body>
    </html>
    """


    
    try:
        # Enviar solo el texto, sin PDF
        mensaje = MIMEText(cuerpo, 'html')
        mensaje['From'] = "billetajevmt@gmail.com"
        mensaje['To'] = destinatario
        mensaje['Subject'] = asunto

        # Agregar copia a vmtdeveloperpy@gmail.com
        mensaje['Cc'] = "vmtdeveloperpy@gmail.com"
        
        contexto = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=contexto) as servidor:
            servidor.login("billetajevmt@gmail.com", "qlju dhxo jbon exlg")
            # Enviar a destinatario principal y copia
            destinatarios = [destinatario, "vmtdeveloperpy@gmail.com"]
            servidor.sendmail("billetajevmt@gmail.com", destinatarios, mensaje.as_string())
        print(f"Correo enviado a {destinatario} con copia a vmtdeveloperpy@gmail.com")
        return True
    except Exception as e:
        print(f"Error al enviar notificación de resumen a {destinatario}: {e}")
        return False

def notificar_incumplimientos_resumen_periodo(resumen_eots, resultado_final, fecha_inicio, fecha_fin):
    """Envía notificaciones de resumen de período a todas las empresas con incumplimientos"""
    notificaciones_enviadas = 0
    notificaciones_fallidas = 0
    
    print(f"\n=== ENVIANDO NOTIFICACIONES DE RESUMEN DE PERÍODO ===")
    print(f"Período: {fecha_inicio} a {fecha_fin}")
    print(f"📊 Total de EOTs en resumen: {len(resumen_eots)}")
    print(f"📊 Total de registros en resultado_final: {len(resultado_final)}")
    
    for nombre_eot in resumen_eots.index:
        try:
            # Obtener email de la empresa (usando el email de prueba por ahora)
            email = resumen_eots.loc[nombre_eot]['Email']
            print(f"Email de {nombre_eot}: {email}")
            
            # Obtener fechas específicas de infracción para esta empresa
            fechas_infraccion = resultado_final[resultado_final['Nombre EOT'] == nombre_eot]['Fecha'].tolist()
            fechas_infraccion_str = formatear_fechas_para_tabla(fechas_infraccion)
            
            print(f"Enviando resumen de período a {nombre_eot} ({email})...")
            print(f"  Fechas con infracciones: {fechas_infraccion_str}")
            
            if enviar_notificacion_resumen_periodo(
                destinatario=email,
                nombre_eot=nombre_eot,
                resumen_eot=resumen_eots.loc[nombre_eot].to_dict(),
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                fechas_infraccion=fechas_infraccion_str
            ):
                notificaciones_enviadas += 1
                print(f"  ✅ Email enviado exitosamente a {email}")
            else:
                notificaciones_fallidas += 1
                print(f"  ❌ Error al enviar email a {email}")
        except Exception as e:
            print(f"  ❌ Error procesando {nombre_eot}: {e}")
            notificaciones_fallidas += 1
    
    print(f"\n📊 RESUMEN DE NOTIFICACIONES DE PERÍODO:")
    print(f"✅ Notificaciones enviadas exitosamente: {notificaciones_enviadas}")
    print(f"❌ Notificaciones fallidas: {notificaciones_fallidas}")
    print(f"📧 Total de empresas notificadas: {notificaciones_enviadas + notificaciones_fallidas}")
    
    return notificaciones_enviadas, notificaciones_fallidas

# --- función para exportar el resumen a Excel ---
def exportar_resumen_excel(resumen_eots, nombre_archivo):
    """Exporta el resumen de EOTs a un archivo Excel"""
    try:
        # Crear un DataFrame con el resumen
        df_export = resumen_eots.copy()
        df_export.reset_index(inplace=True)
        df_export.rename(columns={'index': 'Nombre EOT'}, inplace=True)
        
        # Reordenar columnas para mejor presentación
        columnas_ordenadas = ['Nombre EOT', 'Días con <100%', 'Promedio % Ruta', 'Buses Reportado', 'Buses en Ruta', 'Email', 'Fechas con problemas']
        df_export = df_export[columnas_ordenadas]
        
        # Crear el archivo Excel
        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name='Resumen EOTs', index=False)
            
            # Obtener la hoja de trabajo para formatear
            worksheet = writer.sheets['Resumen EOTs']
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Formatear encabezados
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Formatear celdas de porcentaje
            for row in range(2, worksheet.max_row + 1):
                # Columna de porcentaje (índice 2)
                cell = worksheet.cell(row=row, column=3)
                cell.number_format = '0.00"%"'
                cell.alignment = center_alignment
                
                # Columna de días (índice 2)
                cell_dias = worksheet.cell(row=row, column=2)
                cell_dias.alignment = center_alignment
                
                # Columna de buses (índices 4 y 5)
                for col in [4, 5]:
                    cell_buses = worksheet.cell(row=row, column=col)
                    cell_buses.number_format = '0'
                    cell_buses.alignment = center_alignment
        
        print(f"✅ Archivo Excel generado: {nombre_archivo}")
        return True
        
    except Exception as e:
        print(f"❌ Error al generar archivo Excel: {e}")
        return False

hostCID = "168.90.177.232"
databaseCID = "bbdd-monitoreo-cid"
userCID = "cid_admin_user"
passwordCID = "vmtdmtcidccm"
portCID = "2024"

hostMON = "monitoreo.vmt.gov.py"
databaseMON = "bbdd-monitoreo-prod"
userMON = "jefe-CID"
passwordMON = "vmtdmt"
portMON = "5432"

hostZURE = "flex-opentransit-prod-1.postgres.database.azure.com"
databaseZURE = "opentransit-prod"
userZURE = "vmtuser"
passwordZURE = "$$Tr4nsp0rt3$$"
portZURE = "5432"


# Fechas del período
#yesterday = date.today() - timedelta(days=1)
#yesterday_str = yesterday.strftime('%Y-%m-%d')

# Lógica de fechas dinámicas basada en el día del mes
fecha_actual = datetime.now()
dia_mes = fecha_actual.day

if 13 <= dia_mes <= 20:
    # Si está entre el día 12 y 20, consultar del 1 al 15 del mes actual
    desde = fecha_actual.replace(day=1)
    hasta = fecha_actual.replace(day=15)
elif 1 <= dia_mes <= 12:
    # Si está entre el día 1 y 4, consultar del 16 al último día del mes anterior
    mes_anterior = fecha_actual.replace(day=1) - timedelta(days=1)
    desde = mes_anterior.replace(day=16)
    hasta = mes_anterior
else:
    # En cualquier otro caso, no ejecutar
    desde = None
    hasta = None
    sw = False  # Deshabilitar ejecución

# Función para generar fechas diarias
def generar_fechas_diarias(desde, hasta):
    fechas = []
    fecha_actual = desde
    while fecha_actual <= hasta:
        start = fecha_actual.replace(hour=0, minute=0, second=0)
        end = fecha_actual.replace(hour=23, minute=59, second=59)
        fechas.append((start, end))
        fecha_actual += timedelta(days=1)
    return fechas


# CONSULTAR SI SE QUIERE AVANZAR
if desde is None or hasta is None:
    print("No corresponde ejecutar el procesamiento en esta fecha.")
    sys.exit()

# Generar lista de fechas diarias
fechas_diarias = generar_fechas_diarias(desde, hasta)
print(f"Consultando {len(fechas_diarias)} días desde {desde.strftime('%Y-%m-%d')} hasta {hasta.strftime('%Y-%m-%d')}")

# Lista para almacenar resultados de todos los días
resultados_todos_dias = []

# Obtener datos de EOTs permisionarios
connCID = None
cur = None
try:
    connCID = psycopg2.connect(host=hostCID, database=databaseCID, user=userCID, password=passwordCID, port=portCID)
    cur = connCID.cursor()
    cur.execute("SELECT eot_nombre, cod_catalogo, id_eot_vmt_hex, e_mail FROM eots WHERE permisionario = TRUE")
    rows = cur.fetchall()
    colnames = [desc[0] for desc in cur.description]
    datasetEOT = pd.DataFrame(rows, columns=colnames)
    datasetEOT.reset_index(inplace=True)
    datasetEOT.rename(columns={'index': 'N°'}, inplace=True)
    datasetEOT['N°'] += 1  # Para que empiece en 1 en vez de 0
    # Reordenar columnas para que 'N°' sea la primera
    cols = datasetEOT.columns.tolist()
    cols = [cols[-1]] + cols[:-1]
    datasetEOT = datasetEOT[cols]
    # Renombrar encabezados a español
    datasetEOT.rename(columns={
        'N°': 'N°',
        'eot_nombre': 'Nombre EOT',
        'id_eot_vmt_hex': 'ID EOT HEX',
        'cod_catalogo': 'Cod Catalogo',
        'e_mail': 'Email'
    }, inplace=True)
    datasetEOT = datasetEOT[['N°', 'Nombre EOT', 'ID EOT HEX', 'Cod Catalogo', 'Email']]
    print("Se ejecutó la consulta de los EOTs con los datos de los permisos de los EOTs y se guardó en el datasetEOT")
    
    # Procesar cada día individualmente
    for fecha_idx, (start_date, end_date) in enumerate(fechas_diarias):
        print(f"\n=== PROCESANDO DÍA {fecha_idx + 1}/{len(fechas_diarias)}: {start_date.strftime('%Y-%m-%d')} ===")
        
        # Crear una copia del datasetEOT para este día
        datasetEOT_dia = datasetEOT.copy()
        
        # Agregar columna 'Buses Reportado'
        buses_reportado = []
        # Conexión a la base de datos MON
        try:
            connMON = psycopg2.connect(
                host=hostMON,
                database=databaseMON,
                user=userMON,
                password=passwordMON,
                port=portMON
            )
            curMON = connMON.cursor()
            for idx, row in datasetEOT_dia.iterrows():
                agency_id = row['ID EOT HEX']
                query = """
                    SELECT COUNT(DISTINCT mean_id)
                    FROM app_monitoreo_mensajeoperativo
                    WHERE agency_id = %s AND fecha_hora BETWEEN %s AND %s
                """
                start = start_date.strftime('%Y-%m-%d %H:%M:%S')
                end = end_date.strftime('%Y-%m-%d %H:%M:%S')
                #print("QUERY:", query)
                #print("PARAMS:", (agency_id, start, end))
                curMON.execute(query, (agency_id, start, end))
                result = curMON.fetchone()
                count = result[0] if result is not None else 0
                buses_reportado.append(count)
            curMON.close()
            connMON.close()
        except (Exception, psycopg2.Error) as error:
            print("Error al consultar buses reportados:", error)
            buses_reportado = [0] * len(datasetEOT_dia)

        datasetEOT_dia['Buses Reportado'] = buses_reportado
        print("Se ejecutó la consulta de los Buses Reportado y se guardó en el datasetEOT")

        # Unificar consultas de ID SAM y ID RUTA en un solo recorrido
        # data_idsam = []
        data_idruta = []
        try:
            connCID2 = psycopg2.connect(host=hostCID, database=databaseCID, user=userCID, password=passwordCID, port=portCID)
            curCID2 = connCID2.cursor()
            for idx, row in datasetEOT_dia.iterrows():
                cod_catalogo = row['Cod Catalogo']
                # Consulta para ID RUTA
                query_ruta = """
                    SELECT DISTINCT ruta_hex
                    FROM catalogo_rutas
                    WHERE id_eot_catalogo = %s
                """
                #print("QUERY RUTA:", query_ruta)
                #print("PARAMS RUTA:", (cod_catalogo,))
                curCID2.execute(query_ruta, (cod_catalogo,))
                results_ruta = curCID2.fetchall()
                for (ruta_hex,) in results_ruta:
                    data_idruta.append({
                        'Cod Catalogo': cod_catalogo,
                        'ID RUTA': ruta_hex
                    })
            curCID2.close()
            connCID2.close()
        except (Exception, psycopg2.Error) as error:
            print("Error al consultar declaracion_jurada o catalogo_rutas:", error)
        print("Se ejecutó la consulta de los ID RUTA y se guardó en el datasetIDRUTA")
        datasetIDRUTA = pd.DataFrame(data_idruta, columns=['Cod Catalogo', 'ID RUTA'])

        # Unificar el cálculo de 'Buses operando' y 'Buses operando ruta' en un solo ciclo
        buses_operando = []
        buses_operando_ruta = []
        try:
            connZURE = psycopg2.connect(host=hostZURE, database=databaseZURE, user=userZURE, password=passwordZURE, port=portZURE)
            curZURE = connZURE.cursor()
            for idx, row in datasetEOT_dia.iterrows():
                cod_catalogo = row['Cod Catalogo']
                # IDs RUTA para este catálogo
                idruta_list = datasetIDRUTA[datasetIDRUTA['Cod Catalogo'] == cod_catalogo]['ID RUTA'].tolist()
                start = start_date.strftime('%Y-%m-%d %H:%M:%S')
                end = end_date.strftime('%Y-%m-%d %H:%M:%S')

                # Buses operando ruta (por ID RUTA)
                if idruta_list:
                    query_ruta = """
                        SELECT COUNT(DISTINCT idsam)
                        FROM c_transacciones
                        WHERE fechahoraevento BETWEEN %s AND %s AND idrutaestacion IN %s
                    """
                    #print("QUERY ZURE RUTA:", query_ruta)
                    #print("PARAMS ZURE RUTA:", (start, end, tuple(idruta_list)))
                    curZURE.execute(query_ruta, (start, end, tuple(idruta_list)))
                    result = curZURE.fetchone()
                    count = result[0] if result is not None else 0
                    buses_operando_ruta.append(count)
                else:
                    buses_operando_ruta.append(0)
            curZURE.close()
            connZURE.close()
        except (Exception, psycopg2.Error) as error:
            print("Error al consultar c_transacciones en ZURE:", error)
            buses_operando = [0] * len(datasetEOT_dia)
            buses_operando_ruta = [0] * len(datasetEOT_dia)
        print("Se ejecutó la consulta de los Buses operando ruta y se guardó en el datasetEOT")
        datasetEOT_dia['Buses operando en ruta'] = buses_operando_ruta
        
        # Agregar columna 'Porcentaje' a datasetEOT
        def calcular_porcentaje(operando, reportado):
            if reportado and reportado > 0:
                return f"{(operando / reportado) * 100:.2f}%"
            elif operando == 0:
                return "100.00%"
            else:
                return "0.00%"

        datasetEOT_dia['Porcentaje en ruta'] = datasetEOT_dia.apply(lambda row: calcular_porcentaje(row['Buses Reportado'], row['Buses operando en ruta']), axis=1)
        
        # Agregar columna de fecha
        datasetEOT_dia['Fecha'] = start_date.strftime('%Y-%m-%d')
        
        # Filtrar solo EOTs con porcentaje en ruta < 100%
        def extraer_porcentaje(porcentaje_str):
            try:
                return float(porcentaje_str.replace('%', ''))
            except:
                return 0.0
        
        datasetEOT_dia['Porcentaje_Num'] = datasetEOT_dia['Porcentaje en ruta'].apply(extraer_porcentaje)
        eots_bajo_100 = datasetEOT_dia[datasetEOT_dia['Porcentaje_Num'] < 100.0].copy()
        
        if not eots_bajo_100.empty:
            print(f"\nEOTs con porcentaje en ruta < 100% para {start_date.strftime('%Y-%m-%d')}:" )
            print(tabulate(eots_bajo_100[['Nombre EOT', 'Buses Reportado', 'Buses operando en ruta', 'Porcentaje en ruta', 'Fecha']], 
                          headers="keys", tablefmt='psql', showindex=False))
            
            resultados_todos_dias.append(eots_bajo_100)
        else:
            print(f"No hay EOTs con porcentaje en ruta < 100% para {start_date.strftime('%Y-%m-%d')}")
        
        print("Se ejecutó la consulta de los Porcentaje en ruta y se guardó en el datasetEOT")

    print(f"\n🔍 DEBUG: Terminó el bucle de fechas. Total de días procesados: {len(resultados_todos_dias)}")

    # Mostrar resumen final
    if resultados_todos_dias:
        print("\n🔍 DEBUG: Entrando al resumen final...")
        print("\n" + "="*80)
        print("RESUMEN FINAL - EOTs CON PORCENTAJE EN RUTA < 100%")
        print("="*80)
        
        # Combinar todos los resultados
        resultado_final = pd.concat(resultados_todos_dias, ignore_index=True)
        
        # CREAR TABLA DE DETALLES SI NO EXISTE
        print(f"\n🏗️  CREANDO TABLA DE DETALLES DE INCUMPLIMIENTOS...")
        if crear_tabla_detalles_incumplimientos():
            print(f"✅ Tabla de detalles lista")
        else:
            print(f"❌ Error al crear tabla de detalles")
        
        # GUARDAR DETALLES INDIVIDUALES DE INCUMPLIMIENTOS
        print(f"\n💾 GUARDANDO DETALLES INDIVIDUALES DE INCUMPLIMIENTOS...")
        print(f"📊 Total de registros de incumplimientos: {len(resultado_final)}")
        print(f"📅 Período: {desde.strftime('%Y-%m-%d')} a {hasta.strftime('%Y-%m-%d')}")
        
        if insertar_detalles_incumplimientos(resultado_final):
            print(f"✅ Detalles de incumplimientos guardados exitosamente")
        else:
            print(f"❌ Error al guardar detalles de incumplimientos")
        
        # GUARDAR INCUMPLIMIENTOS AGRUPADOS EN LA BASE DE DATOS (RESULTADO FINAL)
        print(f"\n💾 GUARDANDO INCUMPLIMIENTOS AGRUPADOS EN LA BASE DE DATOS...")
        
        if insertar_incumplimientos(resultado_final, desde, hasta):
            print(f"✅ Incumplimientos agrupados guardados exitosamente en la base de datos")
        else:
            print(f"❌ Error al guardar incumplimientos agrupados en la base de datos")
        
        # Agrupar por EOT y mostrar estadísticas
        resumen_eots = resultado_final.groupby('Nombre EOT').agg({
            'Fecha': ['count', lambda x: ', '.join(sorted(x.unique()))],
            'Porcentaje_Num': 'mean',
            'Buses Reportado': 'mean',
            'Buses operando en ruta': 'mean',
            'Email': 'first'  # Tomar el primer email para cada EOT
        }).round(2)
        
        resumen_eots.columns = ['Días con <100%', 'Fechas con problemas', 'Promedio % Ruta', 'Buses Reportado', 'Buses en Ruta', 'Email']
        resumen_eots = resumen_eots.sort_values('Días con <100%', ascending=False)
        
        print("\nResumen por EOT:")
        print(tabulate(resumen_eots, headers="keys", tablefmt='psql', showindex=True))
        print("\nVisualización clara del resumen generado:")
        print(resumen_eots)
        
        print(f"\nTotal de EOTs con porcentaje < 100%: {len(resumen_eots)}")
        print(f"Total de días con problemas: {len(resultado_final)}")
        
        # Exportar a Excel
        exportar_resumen_excel(resumen_eots, "resumen_eots.xlsx")
        
        # GENERAR NOTIFICACIONES DE INCUMPLIMIENTOS EN LA BASE DE DATOS
        print(f"\n🔔 GENERANDO NOTIFICACIONES DE INCUMPLIMIENTOS...")
        try:
            notificaciones_generadas = generar_notificaciones_incumplimientos()
            if notificaciones_generadas:
                print(f"✅ Se generaron {len(notificaciones_generadas)} notificaciones de incumplimientos")
                for notif in notificaciones_generadas:
                    print(f"  - {notif['eot_nombre']}: {notif['cantidad']} incumplimientos (Notificación #{notif['nro_notificacion']})")
            else:
                print("ℹ️  No se generaron nuevas notificaciones de incumplimientos")
        except Exception as e:
            print(f"❌ Error al generar notificaciones de incumplimientos: {e}")
            import traceback
            traceback.print_exc()

        # ENVIAR NOTIFICACIONES DE RESUMEN DE PERÍODO A LAS EMPRESAS CON INCUMPLIMIENTOS
        print(f"\n🚀 INICIANDO ENVÍO DE NOTIFICACIONES...")
        fecha_inicio_str = desde.strftime('%Y-%m-%d')
        fecha_fin_str = hasta.strftime('%Y-%m-%d')
        print(f"📅 Período: {fecha_inicio_str} a {fecha_fin_str}")
        print(f"📊 EOTs a notificar: {len(resumen_eots)}")
        
        try:
            notificar_incumplimientos_resumen_periodo(resumen_eots, resultado_final, fecha_inicio_str, fecha_fin_str)
            print(f"✅ Proceso de notificaciones completado")
        except Exception as e:
            print(f"❌ Error en el proceso de notificaciones: {e}")
            import traceback
            traceback.print_exc()
        
        # Enviar resumen general por email con PDF a rolandog@mopc.gov.py con copia a lprafael1710@gmail.com, hatoweb@gmail.com, transporte.mopc@gmail.com
        print(f"\n📧 Enviando resumen general con PDF por email...")
        try:
            # Crear mensaje con múltiples destinatarios
            mensaje = MIMEMultipart()
            mensaje['From'] = "billetajevmt@gmail.com"
            mensaje['To'] = "rolandog@mopc.gov.py"
            mensaje['Cc'] = "lprafael1710@gmail.com,hatoweb@gmail.com,transporte.mopc@gmail.com"
            mensaje['Subject'] = f"Resumen EOTs - Período {desde.strftime('%Y-%m-%d')} a {hasta.strftime('%Y-%m-%d')}"
            
            # Generar tabla HTML del resumen
            tabla_html = generar_tabla_resumen_html(resumen_eots)
            
            cuerpo = f"""
<html>
<body style="font-family: Arial, sans-serif; color: #333; line-height: 1.5;">
    <h2 style="color: #004080;">Resumen del Monitoreo de EOTs</h2>
    <p><strong>Período:</strong> {desde.strftime('%Y-%m-%d')} a {hasta.strftime('%Y-%m-%d')}</p>
    
    <h3 style="color: #800000;">Estadísticas Generales:</h3>
    <ul>
        <li><strong>Total de EOTs con incumplimientos:</strong> {len(resumen_eots)}</li>
        <li><strong>Total de días con problemas:</strong> {len(resultado_final)}</li>
        <li><strong>Notificaciones enviadas:</strong> Se enviaron notificaciones de resumen de período a todas las empresas con incumplimientos</li>
    </ul>
    
    <h3 style="color: #800000;">Detalle por EOT:</h3>
    {tabla_html}
    
    <p><em>Adjunto encontrará el reporte detallado en Excel.</em></p>
    
    <p>Este reporte fue generado automáticamente por el <strong>Sistema de Monitoreo VMT</strong>.</p>
</body>
</html>
"""
            
            mensaje.attach(MIMEText(cuerpo, 'html'))
            
            # Adjuntar Excel
            with open("resumen_eots.xlsx", "rb") as adjunto:
                parte = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                parte.set_payload(adjunto.read())
                encoders.encode_base64(parte)
                parte.add_header('Content-Disposition', 'attachment; filename=resumen_eots.xlsx')
                mensaje.attach(parte)
            
            # Lista de destinatarios (principal + copia)
            destinatarios = ["rolandog@mopc.gov.py", "lprafael1710@gmail.com", "hatoweb@gmail.com", "transporte.mopc@gmail.com"]
            
            contexto = ssl.create_default_context()
            with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=contexto) as servidor:
                servidor.login("billetajevmt@gmail.com", "qlju dhxo jbon exlg")
                servidor.sendmail("billetajevmt@gmail.com", destinatarios, mensaje.as_string())
            
            print("✅ Resumen general con Excel enviado exitosamente a rolandog@mopc.gov.py (copia a lprafael1710@gmail.com, hatoweb@gmail.com, transporte.mopc@gmail.com)")
        except Exception as e:
            print(f"❌ Error al enviar resumen general: {e}")
        
    else:
        print("\nNo se encontraron EOTs con porcentaje en ruta < 100% en ningún día del período.")

except (Exception, psycopg2.Error) as error:
    print("Error al conectar a PostgreSQL", error)
finally:
    if connCID:
        try:
            cur.close()
        except:
            pass
        connCID.close()
        print("Conexión a la base de datos cerrada correctamente")

# Ejemplo de uso (comentado para evitar envío automático):
# enviar_email_con_adjunto(
#      destinatario="hatoweb@gmail.com",
#      asunto="Resumen EOTs",
#      cuerpo="Adjunto el resumen de EOTs con porcentaje en ruta < 100%.",
#      archivo_adjunto="resumen_eots.pdf",
#      remitente="billetajevmt@gmail.com",
#      password="qlju dhxo jbon exlg",
#      servidor_smtp="smtp.gmail.com",
#      puerto_smtp=465
#  )

# Ejemplos de consulta de detalles de incumplimientos:
# 
# # Consultar todos los detalles
# detalles = consultar_detalles_incumplimientos()
# print(detalles)
# 
# # Consultar por rango de fechas
# detalles_fecha = consultar_detalles_incumplimientos(
#     fecha_inicio='2025-09-01', 
#     fecha_fin='2025-09-05'
# )
# print(detalles_fecha)
# 
# # Consultar por EOT específico
# detalles_eot = consultar_detalles_incumplimientos(eot_nombre='DE LA CONQUISTA')
# print(detalles_eot)




